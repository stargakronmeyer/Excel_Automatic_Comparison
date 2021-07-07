# -*- coding: utf-8 -*-
"""
Created on Wed Jul  7 08:15:48 2021

@author: STKRO

1 - Ideia do Script: Comparar planilhas de excel e ver se elas são iguais.
    If the same, 
        return None and prints on screen "Tables are equal"
    else:
        returns the number os differences and their location (column and row)
v1
"""
import xlsxwriter
import pandas as pd 
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill

def sheets_read():
    """
    Parameters
    ----------
    Nothing , no input
    
    Returns
    -------
    df_1,df_2: returns the dfs that are read from the excel files and also their names
        
    """
    name_1=r"E:\stkro\Documents\Tarefas.xlsx"
    name_2=r"E:\stkro\Documents\Tasks.xlsx"
    #path to the excel spreadsheet
    df_1=pd.read_excel(name_1)
    df_2=pd.read_excel(name_2)
    return df_1,df_2,name_1,name_2

def NanToNone(df_1,df_2):
    """
    Parameters
    ----------
    df_1,df_2: Dataframes are from the excel
    -------
    Returns
    df_1,df_2: returns the dfs that are read from the excel files and have the same sizes with all the NaN values transformed to None
        
    
    """
    #based on https://stackoverflow.com/questions/14162723/replacing-pandas-or-numpy-nan-with-a-none-to-use-with-mysqldb
    #The ideia here is to replace witha  constant (None - string, as the value None is interpreted by pandas as NaN)
    df_1=df_1.replace({np.nan: 'None'})
    df_2=df_2.replace({np.nan: 'None'})
    
    if df_1.shape==df_2.shape:
        pass
    else:
        difference_row=df_1.shape[0]-df_2.shape[0]
        abs_difference=abs(difference_row)
        if difference_row<0:
            #df_2 has more rows than df_1, so we add rows on df_1
            for i in range(abs_difference):
                df_1.loc[len(df_1)]=['None']*len(df_1.columns)
        else:
            for i in range(abs_difference):
                df_2.loc[len(df_2)]=['None']*len(df_2.columns)
        difference_column=df_1.shape[1]-df_2.shape[1]
        abs_difference=abs(difference_column)
        if difference_column<0:
            #df_2 has more rows than df_1, so we add rows on df_1
            for i in range(abs_difference):
                df_1[i]=['None']*len(df_1)
        else:
            for i in range(abs_difference):
                df_2[i]=['None']*len(df_2)
    
    return df_1,df_2

def Compare(df_1,df_2):
    """
    Parameters
    ----------
    df_1,df_2: Dataframes of same size
    -------
    
    Returns
    df_index: df with the 'adress' of the cells that are different on the 2 spreadsheets
    
    """
    #take the mask of the Null Values
    #if those mask are the same the same, we have the same values for None on the two columns. 
    
    Equity=(df_1!=df_2)
    
    columns=Equity.columns
    row=[]
    colss=[]
    
    for col in range(len(columns)):
        #print(columns[col])
        
        if Equity.index[Equity[columns[col]]==True].tolist()==None or (Equity.index[Equity[columns[col]]==True].tolist())==[]:
            #print('oi')
            pass
        else:  
            row.append(Equity.index[Equity[columns[col]]==True].tolist())
            #print(row)
            cols=[col]*len(Equity.index[Equity[columns[col]]==True].tolist())
            colss.append(cols)
    #nested list comprehension
    #this +2 and +1 is important so that the number matches the excel row and not the DF row
    #
    rows=[x + 2  for sublist in row for x in sublist]
    cols=[x + 1 for sublist in colss for x in sublist]
    
    df_index=pd.DataFrame({'Row':rows,'Col':cols})
    
    return df_index

def ColorSpreadsheet(df,name_1,name_2):
    
    workbook1 = openpyxl.load_workbook(name_1)
    workbook2 = openpyxl.load_workbook(name_2)
    
    wb1=workbook1.active
    wb2=workbook2.active
    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
    
    for i in range(len(df)):
        #paints it red
        wb1.cell(df['Row'].iloc[i],df['Col'].iloc[i]).fill=redFill
        wb2.cell(df['Row'].iloc[i],df['Col'].iloc[i]).fill=redFill
    #print(workbook1)
    workbook1.save(name_1)
    workbook2.save(name_2)

def FormatCell(df):
    #https://stackoverflow.com/questions/47179026/convert-number-to-alphabet-that-corresponds-to-excel-column-alphabet
    alphabet=[]
    for i in range(len(df)):
        letter=xlsxwriter.utility.xl_col_to_name(df['Col'].iloc[i]-1)
        adress=letter+str(df['Row'].iloc[i])
        alphabet.append(adress)
    df['Cell']=alphabet
    return df

def main():
    a,b,name_1,name_2=sheets_read()
    c,d=NanToNone(a,b)
    df=Compare(c,d)
    ColorSpreadsheet(df,name_1,name_2)
    #print(df)
    cell_names=FormatCell(df)
    cell_names.to_excel("Index_and_Rows_of_difference.xlsx")
    print(cell_names)
    return cell_names

if __name__ == '__main__':
     main()

    
    